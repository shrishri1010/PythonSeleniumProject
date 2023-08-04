import time
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl

#install openPyExcel  --->> File -- settings --- python inperpreter  ----- install openPyExcel package
#Reading data from Excel
path = "D:\data1.xlsx"
#load the workbook
workbook=openpyxl.load_workbook(path)

#fetch sheetName from workbook
sheet=workbook.get_sheet_by_name("Sheet1")
#sheet=workbook.active

#get rows and coloumns
rows=sheet.max_row      #5
cols=sheet.max_column   #4

print(rows)
print(cols)

for r in range(1,rows+1):       #for loop for rows iteration - 1 is for first row
    for c in range(1,cols+1):    #for loop for coloumns
        print(sheet.cell(row=r,column=c).value,end="   ")
    print()
    

