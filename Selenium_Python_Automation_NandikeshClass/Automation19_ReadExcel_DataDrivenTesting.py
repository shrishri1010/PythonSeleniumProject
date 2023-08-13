import openpyxl

#Data driven testing
#install openPyExcel  --->>
# File -- settings --- python interpreter  ----- install openPyExcel package
# Reading data from Excel
path = r"C:\Users\VEN-KulalSR\Desktop\data1.xlsx"
#load the workbook
workbook=openpyxl.load_workbook(path)

#fetch sheetName from workbook
sheet=workbook.get_sheet_by_name("Sheet1")
#sheet=workbook.active

#get rows and coloumns
rows=sheet.max_row      #5
cols=sheet.max_column   #4

print(rows)
print(cols)

for r in range(1,rows+1):       #for loop for rows iteration - 1 is for first row
    for c in range(1,cols+1):    #for loop for coloumns
        print(sheet.cell(row=r,column=c).value,end="   ")
    print()
    

